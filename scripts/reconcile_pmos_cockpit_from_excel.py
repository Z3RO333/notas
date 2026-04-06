from __future__ import annotations

import argparse
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import requests


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "ordens_sap.xlsx"
PAGE_SIZE = 1000
UPSERT_BATCH_SIZE = 500
PATCH_BATCH_SIZE = 150
NOTE_BATCH_SIZE = 100
RPC_BATCH_SIZE = 1000


def load_env(path: Path) -> dict[str, str]:
  env: dict[str, str] = {}
  for line in path.read_text(encoding="utf-8").splitlines():
    line = line.strip()
    if not line or line.startswith("#") or "=" not in line:
      continue
    key, _, value = line.partition("=")
    env[key.strip()] = value.strip().strip('"').strip("'")
  return env


env = load_env(REPO_ROOT / ".env")
SUPABASE_URL = env["SUPABASE_URL"].rstrip("/")
SERVICE_KEY = env["SUPABASE_SERVICE_KEY"]
HEADERS = {
  "apikey": SERVICE_KEY,
  "Authorization": f"Bearer {SERVICE_KEY}",
  "Content-Type": "application/json",
}


@dataclass
class ParsedOrders:
  total_unique_codes: int
  window_codes: set[str]
  operational_rows: list[dict]
  finance_rows: list[dict]
  by_year: Counter


def normalize_text(value: object) -> str | None:
  if value is None:
    return None
  text = str(value).strip()
  if not text or text.lower() in {"nan", "none"}:
    return None
  return text


def normalize_code(value: object) -> str | None:
  text = normalize_text(value)
  if text is None:
    return None
  if re.fullmatch(r"\d+\.0", text):
    return text[:-2]
  return text


def normalize_status(value: object) -> str | None:
  text = normalize_text(value)
  if text is None:
    return None
  sem_acento = "".join(
    ch for ch in unicodedata.normalize("NFKD", text)
    if not unicodedata.combining(ch)
  )
  return re.sub(r"[^A-Za-z0-9]+", "_", sem_acento).strip("_").upper() or None


def parse_date_only(value: object) -> str | None:
  if value is None:
    return None
  if isinstance(value, datetime):
    return value.date().isoformat()
  if isinstance(value, date):
    return value.isoformat()
  text = str(value).strip()
  if not text:
    return None
  iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
  if iso_match:
    return iso_match.group(0)
  br_match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", text)
  if br_match:
    return f"{br_match.group(3)}-{br_match.group(2).zfill(2)}-{br_match.group(1).zfill(2)}"
  return None


def parse_number(value: object) -> float:
  if value is None:
    return 0.0
  if isinstance(value, (int, float)):
    return float(value)
  text = str(value).strip()
  if not text:
    return 0.0
  try:
    if re.fullmatch(r"-?\d{1,3}(\.\d{3})*(,\d+)?", text):
      return float(text.replace(".", "").replace(",", "."))
    if re.fullmatch(r"-?\d+,\d+", text):
      return float(text.replace(",", "."))
    return float(text)
  except ValueError:
    return 0.0


def ensure_iso_start_of_day(ymd: str | None) -> str | None:
  if ymd is None:
    return None
  return f"{ymd}T00:00:00+00:00"


def read_orders_from_excel(path: Path, start_date: str, end_date: str) -> ParsedOrders:
  workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
  worksheet = workbook.active
  try:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: index for index, header in enumerate(headers)}

    required = ["Ordem", "Tipo de ordem", "Data de entrada"]
    missing = [name for name in required if name not in header_index]
    if missing:
      raise RuntimeError(f"Colunas obrigatorias ausentes no Excel: {missing}")

    unique_operational: dict[str, dict] = {}
    unique_finance: dict[str, dict] = {}
    window_codes: set[str] = set()
    by_year: Counter = Counter()

    imported_at = datetime.now(timezone.utc).isoformat()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
      ordem_codigo = normalize_code(row[header_index["Ordem"]])
      tipo_ordem = (normalize_text(row[header_index["Tipo de ordem"]]) or "").upper()
      data_entrada = parse_date_only(row[header_index["Data de entrada"]])

      if not ordem_codigo or tipo_ordem != "PMOS" or data_entrada is None:
        continue

      data_modif = parse_date_only(row[header_index.get("Data modif.", -1)]) if "Data modif." in header_index else None
      status_raw = normalize_status(row[header_index.get("Status da ordem", -1)]) if "Status da ordem" in header_index else None
      centro = normalize_text(row[header_index.get("Cen.localiz.", -1)]) if "Cen.localiz." in header_index else None
      denominacao = normalize_text(row[header_index.get("Denominação", -1)]) if "Denominação" in header_index else None
      numero_nota = normalize_code(row[header_index.get("Nota", -1)]) if "Nota" in header_index else None
      texto_breve = normalize_text(row[header_index.get("Texto breve", -1)]) if "Texto breve" in header_index else None
      fornecedor_codigo = normalize_text(row[header_index.get("Fornecedor", -1)]) if "Fornecedor" in header_index else None
      fornecedor_nome = normalize_text(row[header_index.get("Nome Fornecedor CT", -1)]) if "Nome Fornecedor CT" in header_index else None

      operational_row = {
        "ordem_codigo": ordem_codigo,
        "numero_nota": numero_nota,
        "status_ordem_raw": status_raw,
        "tipo_ordem": "PMOS",
        "data_entrada": ensure_iso_start_of_day(data_entrada),
        "ordem_detectada_em": ensure_iso_start_of_day(data_entrada),
        "status_atualizado_em": ensure_iso_start_of_day(data_modif or data_entrada),
        "centro": centro,
        "denominacao_unidade": denominacao,
      }
      unique_operational[ordem_codigo] = operational_row

      finance_row = {
        "ordem_codigo": ordem_codigo,
        "tipo_ordem": "PMOS",
        "numero_nota": numero_nota,
        "data_entrada": data_entrada,
        "inicio_programado": None,
        "denominacao_unidade": denominacao,
        "texto_breve": texto_breve,
        "fornecedor_codigo": fornecedor_codigo,
        "fornecedor_nome": fornecedor_nome,
        "custos_estimados": parse_number(row[header_index.get("Custs.estimados", -1)]) if "Custs.estimados" in header_index else 0.0,
        "custos_totais_materiais": parse_number(row[header_index.get("Custos tot.mat.", -1)]) if "Custos tot.mat." in header_index else 0.0,
        "custos_adicionais": parse_number(row[header_index.get("Custs.adicionais", -1)]) if "Custs.adicionais" in header_index else 0.0,
        "custos_totais_reais": parse_number(row[header_index.get("Cust.tot.reais", -1)]) if "Cust.tot.reais" in header_index else 0.0,
        "source_file_name": path.name,
        "importado_em": imported_at,
      }
      unique_finance[ordem_codigo] = finance_row

      by_year[data_entrada[:4]] += 1
      if start_date <= data_entrada <= end_date:
        window_codes.add(ordem_codigo)

    return ParsedOrders(
      total_unique_codes=len(unique_operational),
      window_codes=window_codes,
      operational_rows=list(unique_operational.values()),
      finance_rows=list(unique_finance.values()),
      by_year=by_year,
    )
  finally:
    workbook.close()


def build_in_filter(values: list[str]) -> str:
  cleaned = [value.strip() for value in values if value and value.strip()]
  return "in.(" + ",".join(cleaned) + ")"


def fetch_pmos_rows_in_window(start_iso: str, end_exclusive_iso: str) -> dict[str, dict]:
  rows: dict[str, dict] = {}
  offset = 0

  while True:
    response = requests.get(
      f"{SUPABASE_URL}/rest/v1/ordens_notas_acompanhamento",
      headers={**HEADERS, "Range": f"{offset}-{offset + PAGE_SIZE - 1}"},
      params={
        "select": "ordem_codigo,data_entrada,ordem_detectada_em,created_at,nota_id,sync_id,tipo_ordem,status_ordem_raw",
        "data_entrada": f"gte.{start_iso}",
        "and": f"(data_entrada.lt.{end_exclusive_iso})",
        "or": "(tipo_ordem.is.null,tipo_ordem.neq.PMPL)",
      },
      timeout=120,
    )
    response.raise_for_status()
    batch = response.json()
    if not batch:
      break

    for row in batch:
      ordem_codigo = normalize_code(row.get("ordem_codigo"))
      if ordem_codigo:
        rows[ordem_codigo] = row

    if len(batch) < PAGE_SIZE:
      break
    offset += PAGE_SIZE

  return rows


def fetch_note_creation_dates(note_ids: set[str]) -> dict[str, str]:
  note_dates: dict[str, str] = {}
  ids = sorted(note_ids)

  for index in range(0, len(ids), NOTE_BATCH_SIZE):
    chunk = ids[index:index + NOTE_BATCH_SIZE]
    response = requests.get(
      f"{SUPABASE_URL}/rest/v1/notas_manutencao",
      headers=HEADERS,
      params={
        "select": "id,data_criacao_sap",
        "id": build_in_filter(chunk),
      },
      timeout=120,
    )
    response.raise_for_status()
    for row in response.json():
      note_id = row.get("id")
      note_date = row.get("data_criacao_sap")
      if note_id and note_date:
        note_dates[str(note_id)] = str(note_date)

  return note_dates


def upsert_rows(table: str, rows: list[dict], on_conflict: str) -> int:
  if not rows:
    return 0

  all_keys = sorted({key for row in rows for key in row})
  normalized_rows = [{key: row.get(key) for key in all_keys} for row in rows]
  sent = 0

  for index in range(0, len(normalized_rows), UPSERT_BATCH_SIZE):
    chunk = normalized_rows[index:index + UPSERT_BATCH_SIZE]
    response = requests.post(
      f"{SUPABASE_URL}/rest/v1/{table}",
      headers={**HEADERS, "Prefer": "resolution=merge-duplicates,return=minimal"},
      params={"on_conflict": on_conflict},
      json=chunk,
      timeout=120,
    )
    response.raise_for_status()
    sent += len(chunk)

  return sent


def call_backfill_from_finance(order_codes: list[str]) -> int:
  updated = 0
  if not order_codes:
    return updated

  for index in range(0, len(order_codes), RPC_BATCH_SIZE):
    chunk = order_codes[index:index + RPC_BATCH_SIZE]
    response = requests.post(
      f"{SUPABASE_URL}/rest/v1/rpc/backfill_data_entrada_from_financeiro",
      headers=HEADERS,
      json={"p_codigos": chunk},
      timeout=120,
    )
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, int):
      updated += payload

  return updated


def select_artificial_candidates(current_rows: dict[str, dict], source_window_codes: set[str]) -> list[str]:
  extras = [row for code, row in current_rows.items() if code not in source_window_codes]
  note_ids = {str(row["nota_id"]) for row in extras if row.get("nota_id")}
  note_dates = fetch_note_creation_dates(note_ids)

  candidates: list[str] = []
  for row in extras:
    ordem_codigo = normalize_code(row.get("ordem_codigo"))
    nota_id = row.get("nota_id")
    data_entrada = str(row.get("data_entrada") or "")[:10]
    ordem_detectada = str(row.get("ordem_detectada_em") or "")[:10]
    created_at = str(row.get("created_at") or "")[:10]
    nota_data_criacao = note_dates.get(str(nota_id)) if nota_id else None

    if not ordem_codigo or not nota_id or not nota_data_criacao:
      continue
    if data_entrada != nota_data_criacao:
      continue
    if created_at != ordem_detectada:
      continue

    candidates.append(ordem_codigo)

  return sorted(set(candidates))


def clear_data_entrada(order_codes: list[str], dry_run: bool) -> int:
  if dry_run or not order_codes:
    return len(order_codes)

  cleared = 0
  for index in range(0, len(order_codes), PATCH_BATCH_SIZE):
    chunk = order_codes[index:index + PATCH_BATCH_SIZE]
    response = requests.patch(
      f"{SUPABASE_URL}/rest/v1/ordens_notas_acompanhamento",
      headers={**HEADERS, "Prefer": "return=minimal"},
      params={"ordem_codigo": build_in_filter(chunk)},
      json={"data_entrada": None},
      timeout=120,
    )
    response.raise_for_status()
    cleared += len(chunk)

  return cleared


def resolve_default_window() -> tuple[str, str]:
  today = datetime.now(timezone.utc).date()
  start = date(today.year, 1, 1)
  return start.isoformat(), today.isoformat()


def main() -> None:
  default_start, default_end = resolve_default_window()

  parser = argparse.ArgumentParser(
    description="Reconcilia o cockpit PMOS com a planilha SAP e limpa data_entrada artificial da 00209."
  )
  parser.add_argument("path", nargs="?", default=str(DEFAULT_XLSX), help="Caminho do arquivo ordens_sap.xlsx")
  parser.add_argument("--start-date", default=default_start, help="Inicio do recorte do cockpit (YYYY-MM-DD)")
  parser.add_argument("--end-date", default=default_end, help="Fim do recorte do cockpit (YYYY-MM-DD)")
  parser.add_argument("--dry-run", action="store_true", help="Mostra o plano sem alterar o Supabase")
  args = parser.parse_args()

  source_path = Path(args.path)
  if not source_path.exists():
    raise FileNotFoundError(f"Arquivo nao encontrado: {source_path}")

  start_date = args.start_date
  end_date = args.end_date
  end_exclusive_date = (datetime.fromisoformat(end_date).date() + timedelta(days=1)).isoformat()
  start_iso = f"{start_date}T00:00:00+00:00"
  end_exclusive_iso = f"{end_exclusive_date}T00:00:00+00:00"

  parsed = read_orders_from_excel(source_path, start_date, end_date)
  before_rows = fetch_pmos_rows_in_window(start_iso, end_exclusive_iso)
  before_count = len(before_rows)
  before_extras = before_count - len(parsed.window_codes & set(before_rows))

  print("Fonte SAP")
  print(f"  Arquivo: {source_path}")
  print(f"  PMOS unicas no arquivo: {parsed.total_unique_codes}")
  print(f"  PMOS no recorte {start_date}..{end_date}: {len(parsed.window_codes)}")
  print(f"  Distribuicao por ano: {dict(sorted(parsed.by_year.items()))}")

  print("\nCockpit antes")
  print(f"  PMOS no painel: {before_count}")
  print(f"  Extras fora da fonte no recorte: {before_extras}")

  if args.dry_run:
    candidates = select_artificial_candidates(before_rows, parsed.window_codes)
    print("\nDry-run")
    print(f"  Upsert financeiro planejado: {len(parsed.finance_rows)}")
    print(f"  Upsert operacional planejado: {len(parsed.operational_rows)}")
    print(f"  Limpeza planejada de data_entrada: {len(candidates)}")
    print(f"  Exemplo de ordens artificiais: {candidates[:20]}")
    return

  finance_upserts = upsert_rows("ordens_financeiro_importado", parsed.finance_rows, "ordem_codigo")
  operational_upserts = upsert_rows("ordens_notas_acompanhamento", parsed.operational_rows, "ordem_codigo")
  backfilled = call_backfill_from_finance([row["ordem_codigo"] for row in parsed.finance_rows])

  refreshed_rows = fetch_pmos_rows_in_window(start_iso, end_exclusive_iso)
  artificial_candidates = select_artificial_candidates(refreshed_rows, parsed.window_codes)
  cleared = clear_data_entrada(artificial_candidates, dry_run=False)

  after_rows = fetch_pmos_rows_in_window(start_iso, end_exclusive_iso)
  after_count = len(after_rows)
  after_source_overlap = len(parsed.window_codes & set(after_rows))
  after_extras = after_count - after_source_overlap

  print("\nExecucao")
  print(f"  Upsert financeiro: {finance_upserts}")
  print(f"  Upsert operacional: {operational_upserts}")
  print(f"  Backfill via fonte financeira: {backfilled}")
  print(f"  data_entrada artificial limpa: {cleared}")

  print("\nCockpit depois")
  print(f"  PMOS no painel: {after_count}")
  print(f"  PMOS da fonte presentes no painel: {after_source_overlap}")
  print(f"  Extras remanescentes no recorte: {after_extras}")

  if after_extras:
    extras = sorted(set(after_rows) - parsed.window_codes)
    print(f"  Exemplo de extras remanescentes: {extras[:20]}")


if __name__ == "__main__":
  main()
